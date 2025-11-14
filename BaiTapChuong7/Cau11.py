from openpyxl import load_workbook, Workbook 
import os

def Menu():
    print("-----QUẢN LÝ NHÂN VIÊN-----")
    print("1. Thêm nhân viên và lưu vào file Excel")
    print("2. Đọc danh sách nhân viên từ file Excel")
    print("3. Sắp xếp nhân viên theo tuổi (và lưu lại)") # Làm rõ chức năng
    print("0. Thoát")
    choice = input("Chọn chức năng: ")
    return choice

FILENAME = 'nhanvien.xlsx'

while True:
    os.system('cls')
    choice = Menu()
    
    if choice == '1':
        # Thêm nhân viên và lưu vào file Excel
        ma = input("Nhập mã nhân viên: ")
        ten = input("Nhập tên nhân viên: ")
        tuoi = int(input("Nhập tuổi nhân viên: "))
        
        try:
            # Sửa lại logic: Luôn dùng openpyxl
            if not os.path.exists(FILENAME):
                # Nếu file chưa tồn tại, tạo mới
                workbook = Workbook() # Dùng Workbook của openpyxl
                worksheet = workbook.active
                worksheet.title = "DanhSachNhanVien"
                # Thiết lập tiêu đề cột (dùng .cell)
                worksheet.cell(row=1, column=1, value='MÃ NHÂN VIÊN') 
                worksheet.cell(row=1, column=2, value='TÊN NHÂN VIÊN') 
                worksheet.cell(row=1, column=3, value='TUỔI') 
                row_to_write = 2 # Dòng tiếp theo để ghi là dòng 2
            else:
                # Nếu file đã tồn tại, tải file
                workbook = load_workbook(FILENAME)
                worksheet = workbook.active
                row_to_write = worksheet.max_row + 1 # Ghi vào dòng cuối + 1
            
            # Ghi dữ liệu nhân viên mới (dùng .cell)
            worksheet.cell(row=row_to_write, column=1, value=ma)
            worksheet.cell(row=row_to_write, column=2, value=ten)
            worksheet.cell(row=row_to_write, column=3, value=tuoi)
            
            # Lưu file (luôn dùng .save() của openpyxl)
            workbook.save(FILENAME)
            
            print(f"Đã thêm nhân viên và lưu vào file {FILENAME}.")
        
        except PermissionError:
            print(f"Lỗi: Không thể lưu file. Có thể file '{FILENAME}' đang được mở.")
        except Exception as e:
            print(f"Đã xảy ra lỗi: {e}")
            
        input("Nhấn Enter để tiếp tục...")
    
    elif choice == '2':
        # Đọc danh sách nhân viên từ file Excel
        if not os.path.exists(FILENAME):
            print(f"Chưa có file {FILENAME}.")
        else:
            try:
                workbook = load_workbook(FILENAME)
                worksheet = workbook.active
                
                print("Danh sách nhân viên:")
                print(f"{'Mã NV':<10} | {'Tên NV':<20} | {'Tuổi':<5}")
                print("-" * 37)
                
                # Bắt đầu từ dòng 2 để bỏ qua tiêu đề
                for row in range(2, worksheet.max_row + 1): 
                    ma = worksheet.cell(row=row, column=1).value
                    ten = worksheet.cell(row=row, column=2).value
                    tuoi = worksheet.cell(row=row, column=3).value
                    
                    # Kiểm tra nếu dòng không rỗng
                    if ma: 
                        print(f"{ma:<10} | {ten:<20} | {tuoi:<5}")
            
            except Exception as e:
                print(f"Đã xảy ra lỗi khi đọc file: {e}")
                
        input("Nhấn Enter để tiếp tục...")
    
    elif choice == '3':
        # Sắp xếp nhân viên theo tuổi và LƯU LẠI
        if not os.path.exists(FILENAME):
            print(f"Chưa có file {FILENAME}.")
        else:
            try:
                workbook = load_workbook(FILENAME)
                worksheet = workbook.active
                
                # Đọc dữ liệu vào danh sách (bỏ qua tiêu đề)
                nhanvien_list = []
                for row in range(2, worksheet.max_row + 1):
                    ma = worksheet.cell(row=row, column=1).value
                    ten = worksheet.cell(row=row, column=2).value
                    tuoi = worksheet.cell(row=row, column=3).value
                    if ma: # Chỉ thêm nếu có dữ liệu
                        nhanvien_list.append([ma, ten, tuoi])
                
                if not nhanvien_list:
                    print("Không có dữ liệu nhân viên để sắp xếp.")
                    input("Nhấn Enter để tiếp tục...")
                    continue

                # Sắp xếp theo tuổi (cột thứ 3, index 2)
                nhanvien_list.sort(key=lambda nv: nv[2])
                
                # Ghi đè dữ liệu đã sắp xếp vào worksheet
                for i, nv in enumerate(nhanvien_list):
                    # Bắt đầu ghi từ dòng 2 (row=i+2)
                    worksheet.cell(row=i+2, column=1, value=nv[0])
                    worksheet.cell(row=i+2, column=2, value=nv[1])
                    worksheet.cell(row=i+2, column=3, value=nv[2])
                
                # Lưu file sau khi sắp xếp
                workbook.save(FILENAME)
                
                print("Danh sách nhân viên sau khi sắp xếp theo tuổi:")
                print(f"{'Mã NV':<10} | {'Tên NV':<20} | {'Tuổi':<5}")
                print("-" * 37)
                for nv in nhanvien_list:
                    print(f"{nv[0]:<10} | {nv[1]:<20} | {nv[2]:<5}")
                
                print(f"\nĐã sắp xếp và lưu lại vào file {FILENAME}.")
            
            except PermissionError:
                print(f"Lỗi: Không thể lưu file. Có thể file '{FILENAME}' đang được mở.")
            except Exception as e:
                print(f"Đã xảy ra lỗi: {e}")
                
        input("Nhấn Enter để tiếp tục...")
        
    elif choice == '0':
        # Thoát
        break
    
    else:
        print("Lựa chọn không hợp lệ. Vui lòng chọn lại.")
        input("Nhấn Enter để tiếp tục...")